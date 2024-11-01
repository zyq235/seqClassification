import os
import torch
from torch.utils.data.dataset import Dataset
import numpy as np
import openpyxl as xl
import json

use_json_label = True

def get_attr_name(ws):
    attrs=[]
    start_attr = ws.cell(1,2).value
    print(start_attr)
    maxcol = ws.max_column
    for j in range(2, maxcol + 1):
        cell = ws.cell(1, j).value
        if cell==start_attr and j!=2:
            break
        attrs.append(cell)
    return attrs

def check_vec(vec):
    res=False
    for i in vec:
        if i:
            res=True
            break
    return res

def deal_excel_file(file_path, sheet_name):
    pid_list=[]
    wb = xl.load_workbook(file_path)
    ws = wb[sheet_name]
    minrow = 2  # 最小行
    maxrow = ws.max_row  # 最大行
    mincol = 2  # 最小列
    maxcol = ws.max_column  # 最大列
    row_num = 0
    data_x=[]
    attr_names = get_attr_name(ws)
    vec_dim = len(attr_names)
    for i in range(minrow, maxrow + 1):
        row = []
        vec=[]
        col_num=0
        pid_list.append(ws.cell(i,1).value)
        for j in range(mincol, maxcol + 1):
            cell = ws.cell(i, j).value
            try:
                cell=float(cell)
            except:
                cell=None
            vec.append(cell)
            if col_num % vec_dim==vec_dim-1:
                if not (None in vec):
                    row.append(vec)
                vec=[]

                #print(cell, end=" ")
            col_num+=1
        if not row:
            continue
        data_x.append(torch.tensor(row, dtype=torch.float))
        row_num += 1

    torch.manual_seed(1)
    #labels = torch.randint(0,4,(1,row_num))
    #labels = torch.reshape(labels,(-1,))

    #data_x = torch.tensor(data_x)
    return data_x,attr_names,pid_list

def get_label(wb):
    ws = wb['临床症状']
    rating = ws["C3"].value
    pid=ws["A1"].value
    pid=pid.split("/")[0]
    print(pid, rating)
    return pid,rating


def deal_label_files(label_path):
    label_dict=dict()
    for file_path in os.listdir(label_path):
        if not file_path.endswith('.xlsx'):
            continue
        print(file_path)
        label_file_path=os.path.join(label_path,file_path)
        wb = xl.load_workbook(label_file_path)
        pid,label=get_label(wb)
        wb.close()
        label_dict[pid]=label
    return label_dict


class TxtDataset(Dataset):
    def __init__(self, data_path,excel_path, sheet_name):
        #self.main_x=corpus.data
        self.excel_path = os.path.join(data_path, excel_path)
        self.label_path = os.path.join(data_path,  "rating")
        self.label = []
        # reading txt file from file
        # fp = open(txt_filepath, 'r')
        # self.txt_filename = [x.strip() for x in fp]
        # fp.close()
        # label_filepath = os.path.join(data_path, label_filename)
        # fp_label = open(label_filepath, 'r')
        # labels = [int(x.strip()) for x in fp_label]
        #fp_label.close()
        self.main_x,self.attr_names,self.pid_list = deal_excel_file(self.excel_path, sheet_name)
        if use_json_label:
            with open(os.path.join(self.label_path,"label.json"),'r') as f:
                label_dict = json.load(f)
        else:
            label_dict = deal_label_files(self.label_path)
            with open(os.path.join(self.label_path, "label.json"), "w") as f:
                json.dump(label_dict, f)
        for i in self.pid_list:
            self.label.append(label_dict[i])

        self.label = torch.tensor([i+3 for i in self.label])




    def __getitem__(self, index):
        label = torch.LongTensor([self.label[index]])
        return self.main_x[index],label

    def __len__(self):
        return len(self.main_x)

    if __name__ == "__main__":
        data_x, labels, attr_name = deal_excel_file('./arrange2.xlsx', "protein")